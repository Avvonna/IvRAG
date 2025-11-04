from contextlib import redirect_stdout
from typing import Generator

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.worksheet import Worksheet
from rapidfuzz import fuzz, process


def find_top_matches(query, choices, top_n, cutoff=None) -> list[tuple[str, float]]:
    """Возвращает топ-N совпадений с их fuzzy score"""
    matches = process.extract(
        query, choices,
        scorer=fuzz.token_set_ratio,
        limit=top_n,
        score_cutoff=cutoff
    )
    return [(m[0], m[1]) for m in matches]

##### ФУНКЦИИ ДЛЯ РЕПОРТОВ #####


def _parse_formatted_question(col: pd.Series):
    """ Функция, разбивающая отформатированный вопрос на составные части """
    pattern = r"^\[(?P<tag>[^\]]+)?\]\s*(?P<q_clean>[^@|]+)(?:\s*@\s*(?P<detail>[^|]+))?(?:\s*\|\s*(?P<option>.+))?$"

    df = col.str.extract(pattern)

    return df

def _get_q_type(
    col: pd.Series,
    out: list[str] = ["type", "option"]
):
    """
    Вспомогательная функция, возвращающая тип вопроса - MIX, SINGLE, MULTI - и другие составные части
    
    Доступные значения out: 'tag', 'q_clean', 'detail', 'option', 'type'
    """

    QS = _parse_formatted_question(col)

    if "type" in out:
        m_det = QS["detail"].notna()
        m_opt = QS["option"].notna()

        # MIX, SINGLE, MULTI
        QS.loc[m_det, "type"] = "MIX"
        QS.loc[~ m_det & ~ m_opt, "type"] = "SINGLE"
        QS.loc[~ m_det & m_opt, "type"] = "MULTI"

    return QS[out]

def _process_one_ext(
    df: pd.DataFrame,
    df_ext: pd.DataFrame,
    q: str, v: str,
    verbose: int
):
    IMPLED = ["TOWNSIZE"]
    if q not in IMPLED:
        return None
    
    df_ext = df_ext[df_ext["name"] == q].copy()
    
    if v not in df_ext["ext_answer"].unique():
        if verbose >= 1:
            print(f"!!! Ответ '{v}' не найден для кастомного вопроса '{q}'")
            print(f"    Доступные ответы: {sorted(df_ext['ext_answer'].unique())}")
        return None
    
    if "year" not in df.columns:
        mapper = df['wave'].cat.categories.to_series().str.split("-").str.get(0).to_dict()
        df['year'] = df['wave'].map(mapper).astype("category")
        if verbose >= 2:
            print("    * Год добавлен в DB")
    
    if q == "TOWNSIZE":
        # Фильтруем вопросы Q4 - Q10
        Q = "Укажите, пожалуйста, город, в котором Вы проживаете постоянно?"
        cats = df["question"].cat.categories
        qs = cats[cats.str.contains(Q, regex=False)]
        m_qs = df["question"].isin(qs)
        
        # Ответы       
        df_ext = df_ext[df_ext["ext_answer"] == v]
        
        keep = pd.merge(df[m_qs], df_ext, on=["year", "answer"], how="inner").index
        cur_mask = pd.Series(False, index=df.index)
        cur_mask.loc[keep] = True
        return cur_mask

def _process_one_filter(
    df: pd.DataFrame,
    df_QS: pd.DataFrame,
    df_ext: pd.DataFrame | None,
    q: str, v: str,
    verbose: int
) -> pd.Series | None:
    """ Вспомогательная функция, обрабатывающая один фильтрующий вопрос """
        
    # Проверка на кастомные фильтры
    if df_ext is not None:
        ext_mask = _process_one_ext(df, df_ext, q, v, verbose)
        if ext_mask is not None:
            return ext_mask

    cur_mask = pd.Series(True, index=df.index)
    
    # Проверка на наличие вопроса в базе данных
    if not df_QS["question"].eq(q).any():
        if verbose >= 1:
            print(f"!!! Вопрос '{q}' не найден в базе данных и кастомных фильтрах")
        return None
    
    cur_mask &= (df["question"] == q)
    
    if not v == "CHECKED":
        ans_mask = df["answer"] == v
        if ans_mask.sum() == 0:
            if verbose >= 1:
                print(f"!!! Значение '{v}' для вопроса '{q}' не найдено в базе данных")
            if verbose >= 2:
                supp_ans_lst = df.loc[df['question'] == q, 'answer'].unique().tolist()
                print(f"    Список возможных ответов на данный вопрос: {supp_ans_lst}")
            return None
    
        cur_mask &= ans_mask
    
    return cur_mask

def _filter_df(
    df: pd.DataFrame,
    df_QS: pd.DataFrame,
    df_ext: pd.DataFrame | None,
    row: pd.Series,
    verbose: int
) -> tuple[pd.DataFrame|None, list]:
    """ Функция, фильтрующая БД по условиям из строки из файла LISTS """
    filter_v = row.loc[[i for i in row.index if i.lower().startswith("filter_ans_")]].dropna()
    treatment = row.loc[[i for i in row.index if i.lower().startswith("treatment_")]].dropna()
    filter_q = row.loc[[i for i in row.index if i.lower().startswith("filter_q_")]].dropna()

    if verbose >= 2:
        print(f"Логика: {treatment.tolist()}")
        print(f"Вопросы для фильтрации: {filter_q.tolist()}")
        print(f"Значения для фильтрации: {filter_v.tolist()}")
        print()

    if not (filter_q.shape[0] == filter_v.shape[0] == treatment.shape[0]):
        raise ValueError(
            "Число вопросов, значений и операций должно совпадать:\n"
            f"вопросов - {filter_q.shape[0]}; значений - {filter_v.shape[0]}; операций - {treatment.shape[0]}"
        )
    
    if filter_q.shape[0] == 0:
        return df, []
    
    problem_flag = False
    mask = pd.Series(True, index=df.index)
    filter_lst = list(zip(filter_q, filter_v, treatment))

    for q, v, t in filter_lst:
        cur_mask = _process_one_filter(df, df_QS, df_ext, q, v, verbose)
        
        # Не return, чтобы пройтись по всем фильтрам, и проверить их
        if cur_mask is None:
            problem_flag = True
            continue
        
        # Обработка логики
        if t == "AND":
            mask &= cur_mask
        elif t == "OR":
            mask |= cur_mask
        else:
            problem_flag = True
            if verbose >= 1:
                print(f"!!! Допустимые значения операций: AND, OR (дано - '{t}')")
            continue

    # Если возникла проблема с фильтрами, то возвращаем None
    if problem_flag:
        return None, []

    # По указанным фильтрам нет респондентов
    if mask.sum() == 0:
        if verbose >= 1:
            print("!!! По указанным фильтрам нет респондентов")
            print()
        return None, []
    
    filtered_uids = df.loc[mask, "respondent_uid"].unique().tolist()
    if verbose >= 2:
        print(f"Количество подходящих респондентов: {len(filtered_uids)}")

    return df.loc[df["respondent_uid"].isin(filtered_uids)], filter_lst

def _create_pivot_for_question(df: pd.DataFrame, qs_df: pd.DataFrame, q_type: str):
    """ Функция для обработки очередного вопроса из листа, возвращает pivot и pivot_pct """
    if q_type == "SINGLE" or q_type == "MIX":
        index = "answer"
    elif q_type == "MULTI":
        index = "option"
    else:
        raise ValueError(f"Неизвестный тип вопроса: {q_type}. Допустимые значения: MIX, SINGLE, MULTI")
    
    df = pd.merge(df, qs_df, on="question", how="left")
    pivot: pd.DataFrame = pd.pivot_table(
        df, index=index, columns="wave", values="respondent_uid", fill_value=0,
        aggfunc="nunique", observed=True # type: ignore
    ) # type: ignore

    pivot.loc["Всего", :] = df.groupby("wave", observed=True)["respondent_uid"].nunique()
    pivot.index.name = "Ответы"

    return pivot, pivot.divide(pivot.loc["Всего", :], axis=1).drop(index="Всего")

def _write_filters_to_excel(
    ws: Worksheet,
    filters_lst: list,
    starting_row: int,
    starting_col: int,
) -> None:
    """ Вспомогательная функция для записи фильтров в Excel """

    filter_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    ws.cell(row=starting_row, column=starting_col).font = Font(bold=True)
    if len(filters_lst) == 0:
        ws.cell(row=starting_row, column=starting_col).value = "Без фильтров" # type: ignore
        return
    ws.cell(row=starting_row, column=starting_col).value = "Фильтры:" # type: ignore

    for i, (q, v, t) in enumerate(filters_lst):
        ws.cell(row=starting_row, column=starting_col + 1 + i*3).value = t
        ws.cell(row=starting_row, column=starting_col + 1 + i*3).border = filter_border
        ws.cell(row=starting_row, column=starting_col + 1 + i*3).alignment = Alignment(horizontal="center")

        ws.cell(row=starting_row, column=starting_col + 2 + i*3).value = q
        ws.cell(row=starting_row, column=starting_col + 2 + i*3).border = filter_border

        ws.cell(row=starting_row, column=starting_col + 3 + i*3).value = v
        ws.cell(row=starting_row, column=starting_col + 3 + i*3).border = filter_border
    return

def _write_pivot_to_excel(
    ws: Worksheet,
    df_gen: Generator,
    starting_row: int,
    starting_col: int,
    format: str|None = None
):
    """
    Вспомогательная функция для записи пивота в Excel и форматирования значений ячеек
    """
    side_border = Border(left=Side(style='thin'), right=Side(style='thin'))
    top_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    bottom_border = Border(left=Side(style='thin'), right=Side(style='thin'), bottom=Side(style='thin'))
    bottom_total_border = Border(left=Side(style='thin'), right=Side(style='thin'), bottom=Side(style='thin'), top=Side(style='thin'))
    top_fill = PatternFill(start_color='A0E720', end_color='A0E720', fill_type='solid')

    for r_idx, r in enumerate(df_gen, start=starting_row):
        
        # Заголовки столбцов
        if r_idx == starting_row:
            for c_idx, c in enumerate(r, start=starting_col):
                ws.cell(row=r_idx, column=c_idx).value = c
                ws.cell(row=r_idx, column=c_idx).font = Font(bold=True)
                ws.cell(row=r_idx, column=c_idx).border = top_border
                ws.cell(row=r_idx, column=c_idx).fill = top_fill
            continue
        
        for c_idx, c in enumerate(r, start=starting_col):
            # Заголовки строк
            if c_idx == starting_col:
                ws.cell(row=r_idx, column=c_idx).value = c
                ws.cell(row=r_idx, column=c_idx).font = Font(bold=True)
                ws.cell(row=r_idx, column=c_idx).alignment = Alignment(vertical="center")
                ws.cell(row=r_idx, column=c_idx).border = side_border
                continue
            
            # Ячейки
            ws.cell(row=r_idx, column=c_idx).value = c
            ws.cell(row=r_idx, column=c_idx).border = side_border
            if format is not None:
                ws.cell(row=r_idx, column=c_idx).number_format = format

    # Нижняя граница
    for c_idx, c in enumerate(r, start=starting_col):
        # Итоговая строка в пивоте
        if format is None:
            ws.cell(row=r_idx, column=c_idx).border = bottom_total_border
            continue

        ws.cell(row=r_idx, column=c_idx).border = bottom_border

    return r_idx

def _write_question_grouped_to_excel(
    ws: Worksheet,
    pivot: pd.DataFrame,
    starting_row: int,
    starting_col: int,
    question_name: str|None = None,
    applied_filters: list|None = None
) -> int:
    """
    Записывает пивоты в Excel:
    - Если задан вопрос, то пишем пивот, как обычно
    - Если вопрос не задан - дана таблица с процентовкой - пишем только ее

    Возвращает последнюю строку пивота (для связи)
    """
    # Ширина столбца с вариантами ответов
    WIDTH = 25

    if pivot is None:
        raise ValueError("pivot не может быть None")

    if (question_name is not None):
        # Ширина столбца с вариантами ответов
        col_name = get_column_letter(starting_col)
        ws.column_dimensions[col_name].width = WIDTH

        # Заголовок вопроса
        ws.cell(row=starting_row, column=starting_col, value=question_name)
        
        # Фильтр
        filters_start_row = starting_row + 2
        if applied_filters is None:
            raise ValueError("applied_filters может быть только пустым списком, если дан вопрос")
        _write_filters_to_excel(ws, applied_filters, filters_start_row, starting_col)
        
        # Настройка pivot'а
        pivot_start_row = filters_start_row + 2
        cell_fmt = None

    else:
        if applied_filters is not None:
            raise ValueError("applied_filters не может быть не None, если вопрос не задан")
        
        # Настройка pivot'а
        pivot_start_row = starting_row + 3
        cell_fmt = FORMAT_PERCENTAGE_00
        
    gen = dataframe_to_rows(pivot.reset_index(col_level=0), index=False, header=True)
    last_row_idx = _write_pivot_to_excel(ws, gen, pivot_start_row, starting_col, format=cell_fmt)
    
    return last_row_idx

def _write_question_stapled_to_excel(
    ws: Worksheet,
    pivot: pd.DataFrame,
    starting_row: int,
    starting_col: int,
    question_name: str|None = None
) -> int:
    """ Записывает пивоты в stapled Excel """

    # Настройка pivot'а
    cell_fmt = FORMAT_PERCENTAGE_00

    pivot = pivot.reset_index(col_level=0)
    pivot.insert(0, question_name, question_name)

    gen = dataframe_to_rows(pivot, index=False, header=True)
    last_row_idx = _write_pivot_to_excel(ws, gen, starting_row, starting_col, format=cell_fmt)
    
    return last_row_idx

def _process_list_row(
    DB: pd.DataFrame,
    df_QS: pd.DataFrame,
    row: pd.Series,
    verbose: int
) -> tuple[pd.DataFrame, pd.DataFrame] | None:
    """ Функция, которая строит пивоты по ряду из LISTS """
    cur_q = row["question"]
    exact = row["EXACT"]
    
    # Выбираем из DB подходящие вопросы
    if pd.notna(exact):
        qs = df_QS[df_QS["question"] == cur_q]
    else:
        qs = df_QS[df_QS["question"].str.contains(cur_q, regex=False)]

    # Если таких вопросов нет в DB, предупреждаем, скипаем вопрос
    n_qs = qs.shape[0]
    if n_qs == 0:
        
        if verbose >= 1:
            print(f"!!! НЕ найдено вопросов для '{cur_q}'")
        return None
    
    if verbose >= 2:
        print(f"{row.name}) Найдено вопросов ({n_qs}) для '{cur_q}':")
        print("    - ", "\n    - ".join(qs["question"].tolist()), sep="")

    # Проверяем типы вопросов. Если вопрос один, то его тип однозначно SINGLE / MIX,
    # иначе - смотрим на тип, который мы получили с помощью функции _get_q_type
    if n_qs == 1:
        if verbose >= 2:
            print("    Используемый тип вопроса - SINGLE")
        q_type = "SINGLE"
    else:
        q_type = qs["type"].iloc[0]

        # Если несколько вопросов типа MIX, логика ломается
        if q_type == "MIX":
            if verbose >= 1:
                print(
                    f"!!! Вопросы типа MIX должны быть разделены: '{cur_q}'\n"
                    "    - ", "\n    - ".join(qs["question"].tolist()), sep=""
                )
            return None
    
    # Если найдены вопросы разных типов - ошибка,
    # нельзя одновременно строить pivot по опциям и ответам
    if qs["type"].nunique() > 1:
        if verbose >= 1:
            print(f"!!! Найденные вопросы имеют разные типы: {qs['type'].unique().tolist()}:")
            for _, q in qs.iterrows():
                print(f"    - {q['type']}: '{q['question']}'")
        return None

    # Фильтруем дополнительно по текущему вопросу из листа
    df_for_cur_q = DB[DB["question"].isin(qs["question"])]

    # Проверка на пустой результат фильтрации
    if df_for_cur_q.empty:
        if verbose >= 1:
            print(f"!!! Нет ответов на вопрос после фильтров: '{cur_q}'")
        return None
    
    # Строим pivot'ы
    pivot, pivot_pct = _create_pivot_for_question(df_for_cur_q, df_QS, q_type)

    return pivot, pivot_pct

def _create_report_grouped(
    DB: pd.DataFrame,
    df_QS: pd.DataFrame,
    lst_df: pd.DataFrame,
    cols: set,
    df_ext: pd.DataFrame | None,
    ws: Worksheet,
    verbose: int
) -> int:
    """ Функция, которая отвечает за создание обычных репортов по LISTS. """

    # Группируем по одинаковым фильтрам, чтобы не пересчитывать каждый раз
    grouping_cols = lst_df.columns.difference(list(cols)).tolist()
    filter_groupings = lst_df.groupby(grouping_cols, dropna=False)

    counter_bad = 0

    # Табличка с занятыми диапазонами. Из-за группировки по условиям фильтрации
    # вопросы могут рассматриваться не последовательно, поэтому храним табличку
    filled_rows = pd.DataFrame(columns=["start", "end"])

    for group_n, (_, group) in enumerate(filter_groupings, start=1):
        if verbose >= 2:
            print("#" * 25)
            print(f"Группа фильтрации: {group_n}")
            print()
        
        # Фильтрация
        filtered_df, applied_filters = _filter_df(DB, df_QS, df_ext, group.iloc[0], verbose)

        if verbose >= 2:
            print("#" * 25)
            print()
        
        # Если фильтрация не удалась, пропускаем группу, все вопросы помечаем как проблемные
        if filtered_df is None:
            counter_bad += group.shape[0]
            continue

        # Перебираем вопросы по группе
        for i in range(group.shape[0]):
            row = group.iloc[i]
            cur_q = row["question"]
            
            pivots = _process_list_row(filtered_df, df_QS, row, verbose)
            if pivots is None:
                counter_bad += 1
                continue

            pivot, pivot_pct = pivots

            # Проверка на налезание таблиц на другие: начальная позиция уже в занятом диапазоне
            overlap_mask: pd.Series = ((filled_rows["start"] <= row["starting_row"]) & (filled_rows["end"] >= row["starting_row"]))
            if overlap_mask.any():
                counter_bad += 1
                if verbose >= 1:
                    print("!!! Вопросы наползают друг на друга:")
                    for q in overlap_mask[overlap_mask].index.tolist():
                        s, e = filled_rows.loc[q]
                        print(f"    - строки [{s}:{e}] - '{q}'")
                    print(f"    - начало на строке {row['starting_row']} - '{row['question']}'")
            
            last_row = _write_question_grouped_to_excel(ws, pivot, row["starting_row"], row["starting_col"], cur_q, applied_filters)
            last_row = _write_question_grouped_to_excel(ws, pivot_pct, last_row, row["starting_col"])

            filled_rows.loc[cur_q] = [row["starting_row"], last_row]
    
    return counter_bad

def _create_report_stapled(
    DB: pd.DataFrame,
    df_QS: pd.DataFrame,
    lst_df: pd.DataFrame,
    ws: Worksheet,
    verbose: int
):
    """ Функция, которая отвечает за создание схлопнутых репортов по LISTS. """

    counter_bad = 0
    last_row = 2
    STARTING_COL = 2

    # Перебираем вопросы
    for _, row in lst_df.iterrows():
        cur_q = row["question"]
        
        pivots = _process_list_row(DB, df_QS, row, verbose)
        if pivots is None:
            counter_bad += 1
            continue
        
        _, pivot_pct = pivots

        last_row = _write_question_stapled_to_excel(ws, pivot_pct, last_row + 1, STARTING_COL, cur_q)
    
    return counter_bad

def _create_reports(
    DB: pd.DataFrame,
    lists_path: str,
    ext_data_path: str | None,
    output_path: str,
    lists_sheets: list[str] | None = None,
    verbose: int = 1
) -> None:
    """ Функция для создания репортов """

    # Необходимые колонки на странице LISTS
    COLS_grouped = {"question", "EXACT", "starting_row", "starting_col"}
    COLS_seq = {"question", "EXACT"}

    # Список вопросов в DB и их типы
    df_QS = DB["question"].cat.categories.to_frame(name="question")
    df_QS[["type", "option"]] = _get_q_type(df_QS["question"])
    
    # Дополнительные данные
    if ext_data_path is not None:
        df_ext = pd.read_parquet(ext_data_path, engine="fastparquet")
    else:
        df_ext = None

    with pd.ExcelFile(lists_path) as xls:
        lists_sheet_names = [str(i) for i in xls.sheet_names if i != "INFO"]
    
    if lists_sheets is not None:
        diff = set(lists_sheets) - set(lists_sheet_names)
        if len(diff) > 0:
            raise ValueError(
                f"В списке листов '{lists_sheets}' есть листы, которых нет в файле '{lists_path}': {sorted(diff)}"
            )
        
        lists_sheet_names = lists_sheets

    wb = Workbook()
    wb.remove(wb.active) # type: ignore

    # Счетчик проблемных вопросов
    total_bad = 0

    for sheet in lists_sheet_names:
        wb.create_sheet(sheet)
        ws = wb[sheet]
        ws.sheet_view.showGridLines = False

        lst_df = pd.read_excel(lists_path, sheet_name=sheet)

        if verbose >= 1:
            print("#" * 50)
            print(f"STARTED: '{sheet}'")
            print("#" * 50)
            print()
        
        # Если есть колонки с позициями вопросов, считаем, что делаем обычный репорт
        if COLS_grouped.issubset(lst_df.columns):
            lst_df = lst_df.dropna(subset=["question"])
            lst_df.loc[:, "starting_row"] = lst_df["starting_row"].astype(int)
            lst_df.loc[:, "starting_col"] = lst_df["starting_col"].astype(int)

            counter_bad = _create_report_grouped(DB, df_QS, lst_df, COLS_grouped, df_ext, ws, verbose)

        # Если их нет - делаем stapled формат
        elif COLS_seq.issubset(lst_df.columns):
            lst_df = lst_df.dropna(subset=["question"])

            counter_bad = _create_report_stapled(DB, df_QS, lst_df, ws, verbose)

        else:
            raise ValueError(f"В таблице отсутствуют необходимые колонки: {COLS_seq - set(lst_df.columns)}")
        
        total_bad += counter_bad

        if verbose >= 1:
            print()
            print("#" * 50)
            print(f"DONE: '{sheet}'")
            print(f"Проблемных вопросов: {counter_bad}")
            print("#" * 50)
            print("\n")

    if verbose >= 1:
        print("\n\n", "#" * 100, sep="")
        print(f"Всего проблемных вопросов: {total_bad}")
        print("#" * 100)
        
    wb.save(output_path)

def create_reports(
    DB: pd.DataFrame,
    lists_path: str,
    ext_data_path: str,
    output_path: str,
    lists_sheets: list[str] | None = None,
    verbose: int = 1,
    logs_path: str | None = None
) -> None:
    if logs_path is not None:
        with open(logs_path, "w", encoding="utf-8") as f, redirect_stdout(f):
            return _create_reports(DB, lists_path, ext_data_path, output_path, lists_sheets, verbose)
    return _create_reports(DB, lists_path, ext_data_path, output_path, lists_sheets, verbose)

def create_recontact_reports(
    DB: pd.DataFrame,
    lists_path: str,
    ext_data_path: str,
    wave_prev_name: str,
    wave_curr_name: str,
    panelist_prev_path: str,
    panelist_curr_path: str,
    output_path: str,
    lists_sheets: list[str] | None = None,
    verbose: int = 1,
    logs_path: str | None = None
) -> None:
    """ Функция, которая делает reconcat report двух указанных волн """
    diff = set([wave_prev_name, wave_curr_name])\
        .difference(DB["wave"].cat.categories.tolist())
    if len(diff) != 0:
        raise KeyError(f"Волн {sorted(diff)} нет в БД")
    
    pls_list = []
    
    for wn, pl_path in zip(
        [wave_prev_name, wave_curr_name],
        [panelist_prev_path, panelist_curr_path]
    ):
        pl = pd.read_excel(pl_path)
        col = f"W{int(wn.split('-')[1])} {wn.split('-')[0]}"
        m = pl.columns.str.contains(col, regex=False)
        cols = pl.columns[m]
        if cols.empty:
            raise KeyError(f"Нет столбца, содержащего '{col}' - {pl_path}")
        if cols.shape[0] > 1:
            raise KeyError(
                f"Столбец, содержащий '{col}' не единственный: {cols.to_list()} - {pl_path}"
            )
        
        COL = "Panelist ID (@_collector_id)"
        
        if pl.columns[0] != COL:
            print(
                f"!!! Название первой колонки отличается от '{COL}': '{pl.columns[0]}'"
            )
        
        pane_col = pl.columns[0]
        ids_col = cols[0]
        
        if verbose >= 1:
            print(f"Из Panelist'а волны '{wn}' была выбрана колонка '{ids_col}'")
        
        pl = pl[[pane_col, ids_col]]
        pl = pl.dropna(subset=[ids_col]).set_index(pane_col)

        pls_list.append(pl)

    pls = pd.merge(
        pls_list[0], pls_list[1],
        how="inner",
        left_index=True, right_index=True
    )
    
    m1 = DB["respondent_id"].isin(pls.iloc[:, 0]) & (DB["wave"] == wave_prev_name)
    m2 = DB["respondent_id"].isin(pls.iloc[:, 1]) & (DB["wave"] == wave_curr_name)
    
    return create_reports(
        DB[m1 | m2],
        lists_path, ext_data_path,
        output_path, lists_sheets, verbose=verbose,
        logs_path=logs_path
    )

def construct_ext_data(
    town_data_path: str,
    output_path: str
):
    COLS = ["year", "wave", "name", "answer", "ext_answer"]
    
    # Население городов
    df = _get_town_data(town_data_path)[COLS]
    
    # Другие операции
    ...
    
    df.to_parquet(output_path, engine="fastparquet", index=False)

def _get_town_data(path: str, name="TOWNSIZE"):
    tdf = pd.read_excel(path)
    tdf = tdf.rename(columns={"Город в БД": "answer"})\
        .dropna(subset=["answer"])\
        .loc[:, "answer":]\
        .melt("answer", var_name="year", value_name="ext_answer")\
        .dropna(subset="ext_answer")
    tdf["ext_answer"] = pd.cut(
        tdf["ext_answer"],
        bins = [0, 100_000, 500_000, 1_000_000, np.inf],
        labels = ["<100", "100 тыс. - 500 тыс.", "500 тыс. - 1 млн.", ">1 млн."]
    )
    tdf["name"] = name
    tdf["wave"] = pd.NA
    
    for col in tdf.columns:
        tdf[col] = tdf[col].astype("string").astype("category")
       
    return tdf
