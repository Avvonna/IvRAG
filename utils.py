import logging
import os
import time
from datetime import datetime
from pathlib import Path
from typing import Any, Callable, Iterable, Literal, Optional, TypeVar

import dotenv
import pandas as pd
from openai import RateLimitError
from openpyxl.styles import Alignment, Font, PatternFill
from rapidfuzz import fuzz, process

logger = logging.getLogger(__name__)

def setup_environment():
    """Загружает переменные окружения"""
    dotenv.load_dotenv()
    
    api_key = os.getenv("OR_API_KEY")
    db_path = os.getenv("DB_PATH")
    
    if not api_key:
        raise ValueError("OR_API_KEY not found in environment")
    if not db_path:
        raise ValueError("DB_PATH not found in environment")
    
    logger.info(f"Environment loaded: DB_PATH={db_path}")
    return api_key, db_path

def setup_logging(
    mode: Literal['CRITICAL', 'ERROR', 'WARNING', 'INFO', 'DEBUG', 'NOTSET'] = 'NOTSET',
    log_dir: str = "logs",
    enable_console: bool = True
):
    os.makedirs(log_dir, exist_ok=True)
    
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")

    run_dir = os.path.join(log_dir, f"run_{timestamp}")
    os.makedirs(run_dir, exist_ok=True)
    log_filepath = os.path.join(run_dir, "log.txt")
    
    # Создаем форматтер
    formatter = logging.Formatter(
        '%(asctime)s | %(levelname)-8s | %(name)-20s | %(message)s',
        datefmt='%Y-%m-%d %H:%M:%S'
    )
    
    # Настраиваем root logger вручную
    root_logger = logging.getLogger()
    root_logger.setLevel(logging.ERROR)
    
    # Очищаем существующие обработчики
    for handler in root_logger.handlers[:]:
        root_logger.removeHandler(handler)
    
    # Файловый обработчик
    file_handler = logging.FileHandler(log_filepath, encoding='utf-8')
    file_handler.setFormatter(formatter)
    root_logger.addHandler(file_handler)
    
    # Консольный обработчик
    if enable_console:
        console_handler = logging.StreamHandler()
        console_handler.setFormatter(formatter)
        root_logger.addHandler(console_handler)
    
    # Наши кастомные логгеры
    allowed_loggers = ['retriever', 'dreamer', 'planner', 'grounder', 'executor']
    for logger_name in allowed_loggers:
        logging.getLogger(logger_name).setLevel(mode)

    print(f"Logging to file: {log_filepath}")
    return run_dir

def load_data(db_path: str, wave_filter: list[str] = ["2025-03"]) -> pd.DataFrame:
    """Загружает данные из parquet файла + фильтр по волнам """
    logger.info(f"Loading data from {db_path}")
    
    df = pd.read_parquet(Path(db_path), engine="fastparquet")
    logger.info(f"Full dataset shape: {df.shape}")
    
    if "wave" in df.columns and wave_filter:
        df = df[df["wave"].isin(wave_filter)]
        logger.info(f"Filtered by waves: {wave_filter}, new shape: {df.shape}")
    
    return df

T = TypeVar('T')

def retry_call(
    fn: Callable[[], T],
    retries: int = 3,
    base_delay: float = 1.0
) -> T:
    """Повтор с откатом (учитывает rate limit и X-RateLimit-Reset)"""
    delay = base_delay
    last_exc: Optional[Exception] = None

    for attempt in range(1, retries + 1):
        try:
            return fn()
        except RateLimitError as e:
            last_exc = e
            wait = delay
            try:
                resp = getattr(e, "response", None)
                ts = resp.headers.get("X-RateLimit-Reset") if resp and hasattr(resp, "headers") else None
                if ts:
                    ts = float(ts) / 1000.0  # мс -> сек
                    wait = max(0.0, ts - time.time())
            except Exception:
                pass

            if attempt < retries:
                logger.warning(f"Rate limit (attempt {attempt}/{retries}); sleep {wait:.1f}s")
                time.sleep(wait)
                delay *= 2
            else:
                raise
        except Exception as e:
            last_exc = e
            if attempt < retries:
                logger.warning(f"{e} (attempt {attempt}/{retries}); sleep {delay:.1f}s")
                time.sleep(delay)
                delay *= 2
            else:
                raise last_exc
    raise

def take_first_n(x: Iterable, n=30):
    x = sorted(x)

    if len(x) <= n:
        return x
    else:
        return x[:n] + ['...']

def get_unique_questions_info(df: pd.DataFrame):   
    qs_ans = df.groupby(["question"], observed=True).agg(
        waves   = ("wave",   set),
        answers = ("answer", set)
    ).reset_index()

    return qs_ans

def find_top_match(query, choices) -> str:
    match_ = process.extract(
        query, choices,
        scorer=fuzz.token_set_ratio,
        limit=1
    )[0]
    return match_[0]

def split_dict_into_chunks(d: dict, n_chunks:int):
    items = list(d.items())
    chunk_size = len(items) // n_chunks
    remainder = len(items) % n_chunks
    
    chunks = []
    start = 0
    
    for i in range(n_chunks):
        end = start + chunk_size + (1 if i < remainder else 0)
        chunks.append(dict(items[start:end]))
        start = end
    
    return chunks

def save_results_to_excel(
    results: dict[str, Any], 
    provenance: dict[str, list[str]] | None = None, 
    filepath: str | None = None
) -> str:
    """
    Сохраняет результаты в Excel.
    Формат:
    Col A (Row N)   : Ключ
    Col A (Row N+1) : История (текст)
    Col B (Row N...) : Значение/Таблица
    """
    
    if not filepath:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filepath = f"results_{timestamp}.xlsx"
    
    logger.info(f"Starting export to Excel: {filepath}")

    try:
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            sheet_name = 'Report'
            pd.DataFrame().to_excel(writer, sheet_name=sheet_name) # Инициализация
            worksheet = writer.sheets[sheet_name]
            
            # --- Стили ---
            # Ключ: жирный, серый фон
            key_font = Font(bold=True, name='Calibri', size=11)
            key_fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
            key_align = Alignment(vertical='top', wrap_text=True)
            
            # История: курсив, поменьше, серый текст
            prov_font = Font(name='Calibri', size=9, italic=True, color="555555")
            prov_align = Alignment(vertical='top', wrap_text=True)
            
            # Общие настройки
            val_align = Alignment(vertical='top')
            worksheet.column_dimensions['A'].width = 40
            worksheet.column_dimensions['B'].width = 25
            
            current_row = 0 # 0-based index для Pandas
            
            for key, value in results.items():
                # 1. Записываем КЛЮЧ (Col A, Row N)
                cell_key = worksheet.cell(row=current_row + 1, column=1, value=str(key))
                cell_key.font = key_font
                cell_key.fill = key_fill
                cell_key.alignment = key_align
                
                # 2. Записываем ИСТОРИЮ (Col A, Row N+1)
                prov_text = "\n".join(provenance.get(key, [])) if provenance else ""
                cell_prov = worksheet.cell(row=current_row + 2, column=1, value=prov_text)
                cell_prov.font = prov_font
                cell_prov.alignment = prov_align
                
                # 3. Записываем ЗНАЧЕНИЕ (Col B, Row N...)
                rows_occupied_val = 0
                
                if isinstance(value, pd.DataFrame):
                    value.to_excel(writer, sheet_name=sheet_name, startrow=current_row, startcol=1)
                    header_height = value.columns.nlevels if hasattr(value.columns, 'nlevels') else 1
                    rows_occupied_val = len(value) + header_height
                else:
                    val_str = str(value) if isinstance(value, (list, dict)) else value
                    cell_val = worksheet.cell(row=current_row + 1, column=2, value=val_str)
                    cell_val.alignment = val_align
                    rows_occupied_val = 1

                # 4. Расчет отступа
                # Слева занято минимум 2 строки (Key + History), справа rows_occupied_val
                block_height = max(rows_occupied_val, 2)
                
                # Сдвигаем курсор + 2 строки разрыва
                current_row += block_height + 2

        logger.info(f"Successfully saved results to {filepath}")
        return filepath

    except Exception as e:
        logger.error(f"Failed to save Excel: {e}", exc_info=True)
        raise